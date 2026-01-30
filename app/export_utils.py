# app/export_utils.py
"""
این فایل فقط یک کار می‌کند:
تبدیل لیست آیتم‌ها به فایل Excel یا PDF و برگرداندن آن به شکل Bytes (برای دانلود).

چرا جدا؟
- main.py شلوغ نشود
- بعداً برای گزارش‌های دیگر هم همین را کپی کنیم
"""

from io import BytesIO
from datetime import datetime

# --- Excel (.xlsx) ---
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# --- PDF (A4) ---
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet

# اگر می‌خواهی فارسی داخل PDF درست نمایش داده شود، از فونت ویندوز استفاده می‌کنیم (Tahoma).
# نکته: reportlab به صورت پیش‌فرض فونت فارسی ندارد.
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont


def _now_stamp() -> str:
    """برای اسم فایل‌ها (مثلاً 20260129_153012)"""
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def items_to_xlsx_bytes(items: list[dict]) -> BytesIO:
    """
    items: لیستی از dict ها مثل:
      {"id": 1, "name": "Oil Filter", "qty": 10, "location": "A-01"}

    خروجی: یک BytesIO آماده دانلود به عنوان .xlsx
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Items"

    # هدرهای اکسل
    headers = ["ID", "Name", "Qty", "Location"]
    ws.append(headers)

    # ردیف‌ها
    for it in items:
        ws.append([
            it.get("id"),
            it.get("name"),
            it.get("qty"),
            it.get("location"),
        ])

    # تنظیم عرض ستون‌ها که قشنگ‌تر شود
    for col_idx, col in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))

        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

    # ذخیره در حافظه (نه روی دیسک)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def items_to_pdf_bytes(items: list[dict], title: str = "Items Report") -> BytesIO:
    """
    خروجی PDF A4 (Portrait).
    اگر جدول خیلی عریض شد، بعداً می‌کنیم Landscape.

    نکته مهم برای فارسی:
    ما فونت Tahoma ویندوز را رجیستر می‌کنیم تا متن فارسی مربع نشود.
    """
    # رجیستر فونت فارسی/عربی ویندوز (Tahoma)
    # اگر خطا داد یعنی مسیر فونت روی سیستم فرق دارد.
    try:
        pdfmetrics.registerFont(TTFont("Tahoma", r"C:\Windows\Fonts\tahoma.ttf"))
        base_font = "Tahoma"
    except Exception:
        # اگر نشد، با فونت پیش‌فرض ادامه می‌دهیم (ممکن است فارسی را درست نشان ندهد)
        base_font = "Helvetica"

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=24, bottomMargin=24, leftMargin=24, rightMargin=24)

    styles = getSampleStyleSheet()
    style_title = styles["Heading2"]
    style_title.fontName = base_font

    story = []
    story.append(Paragraph(title, style_title))
    story.append(Spacer(1, 12))

    # جدول PDF
    data = [["ID", "Name", "Qty", "Location"]]
    for it in items:
        data.append([
            str(it.get("id", "")),
            str(it.get("name", "")),
            str(it.get("qty", "")),
            str(it.get("location", "")),
        ])

    table = Table(data, hAlign="LEFT")

    table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), base_font),
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("PADDING", (0, 0), (-1, -1), 6),
    ]))

    story.append(table)
    doc.build(story)

    buf.seek(0)
    return buf


def make_export_filename(prefix: str, ext: str) -> str:
    """اسم فایل دانلودی"""
    return f"{prefix}_{_now_stamp()}.{ext}"
