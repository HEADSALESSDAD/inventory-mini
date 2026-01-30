# app/exporters.py
"""
هدف این فایل:
- ساخت خروجی Excel (xlsx)
- ساخت خروجی PDF (A4)
این توابع عمومی هستند، یعنی برای همه بخش‌ها (items, suppliers, ...) قابل استفاده‌اند.
"""

from io import BytesIO
from typing import List, Any

# ===== Excel =====
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# ===== PDF =====
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet


def build_excel_xlsx(
    title: str,
    headers: List[str],
    rows: List[List[Any]],
) -> BytesIO:
    """
    خروجی Excel می‌سازد و به صورت BytesIO برمی‌گرداند.
    چرا BytesIO؟ چون مستقیم به کاربر دانلود می‌دهیم، بدون اینکه فایل را دستی روی دیسک ذخیره کنیم.
    """

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]  # اکسل اسم شیت را حداکثر 31 کاراکتر قبول می‌کند

    # ---------- استایل عنوان/هدر ----------
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="DDDDDD")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # ردیف اول: هدرها
    ws.append(headers)

    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # ---------- دیتا ----------
    for r in rows:
        ws.append(r)

    # ---------- Freeze + Filter ----------
    ws.freeze_panes = "A2"           # ردیف هدر ثابت بماند
    ws.auto_filter.ref = ws.dimensions  # فیلتر روی کل دیتا

    # ---------- تنظیم عرض ستون‌ها (تقریبی) ----------
    for col_idx in range(1, len(headers) + 1):
        max_len = 10
        for row_idx in range(1, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

    # خروجی به حافظه
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def build_pdf_a4(
    title: str,
    headers: List[str],
    rows: List[List[Any]],
) -> BytesIO:
    """
    خروجی PDF روی کاغذ A4 می‌سازد و به صورت BytesIO برمی‌گرداند.
    """

    bio = BytesIO()

    # A4 + مارجین‌ها
    doc = SimpleDocTemplate(
        bio,
        pagesize=A4,
        rightMargin=24,
        leftMargin=24,
        topMargin=24,
        bottomMargin=24,
        title=title
    )

    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph(title, styles["Title"]))
    story.append(Spacer(1, 12))

    # جدول: اول هدرها، بعد دیتا
    data = [headers] + rows

    table = Table(data, repeatRows=1)  # repeatRows=1 یعنی هدر در صفحه بعدی هم تکرار شود

    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.black),
        ("GRID",       (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN",      (0, 0), (-1, 0), "CENTER"),
        ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
        ("FONTSIZE",   (0, 0), (-1, -1), 9),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
    ]))

    story.append(table)

    doc.build(story)
    bio.seek(0)
    return bio
