# app/main.py
"""
Inventory Mini - FastAPI + SQLite + Simple Warehouse UI

What this file does (beginner-friendly):
1) Creates the FastAPI app
2) Creates database tables (SQLite) if they don't exist
3) Provides CRUD APIs for /items
4) Serves a clean HTML UI at "/"
5) Exports inventory to Excel (.xlsx) and PDF (A4)

Note:
- UI is English only (as requested).
- UI is rendered as a template, but items are loaded via JavaScript (no "stuck loading").
"""

from __future__ import annotations

from io import BytesIO
from datetime import datetime
from pathlib import Path
from typing import Optional, List

from fastapi import FastAPI, Depends, Request, HTTPException
from fastapi.responses import HTMLResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session

# Your local modules (you already have these files in /app)
from .db import SessionLocal, engine, Base
from . import models, schemas

# -----------------------------
# App + Templates
# -----------------------------

app = FastAPI(title="Inventory Mini", version="0.1.0")

# Always use an absolute path for templates (prevents path issues on Windows)
TEMPLATES_DIR = Path(__file__).resolve().parent / "templates"
templates = Jinja2Templates(directory=str(TEMPLATES_DIR))

# Create tables once at startup (SQLite needs it)
Base.metadata.create_all(bind=engine)


# -----------------------------
# Database session (Dependency)
# -----------------------------
def get_db():
    """
    Creates a DB session for each request and closes it afterwards.
    This avoids leaving connections open.
    """
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


# -----------------------------
# Health
# -----------------------------
@app.get("/health")
def health():
    return {"status": "ok", "app": "Inventory Mini"}


# -----------------------------
# UI (Home)
# -----------------------------
@app.get("/", response_class=HTMLResponse)
def home(request: Request):
    """
    Serves the UI shell.
    Items are loaded by JS from /items (so UI stays fast and never hangs on loading).
    """
    return templates.TemplateResponse("index.html", {"request": request})


# -----------------------------
# CRUD APIs
# -----------------------------
@app.get("/items", response_model=List[schemas.ItemOut])
def list_items(db: Session = Depends(get_db)):
    """
    Returns all items in the database.
    """
    items = db.query(models.Item).order_by(models.Item.id.desc()).all()
    return items


@app.post("/items", response_model=schemas.ItemOut)
def add_item(item: schemas.ItemCreate, db: Session = Depends(get_db)):
    """
    Creates a new item.
    """
    db_item = models.Item(
        name=item.name,
        sku=item.sku,
        qty=item.qty,
        price=item.price,
    )
    db.add(db_item)
    db.commit()
    db.refresh(db_item)
    return db_item


@app.get("/items/{item_id}", response_model=schemas.ItemOut)
def get_one_item(item_id: int, db: Session = Depends(get_db)):
    """
    Returns one item by ID.
    """
    db_item = db.query(models.Item).filter(models.Item.id == item_id).first()
    if not db_item:
        raise HTTPException(status_code=404, detail="Item not found")
    return db_item


@app.put("/items/{item_id}", response_model=schemas.ItemOut)
def update_one_item(item_id: int, item: schemas.ItemCreate, db: Session = Depends(get_db)):
    """
    Updates an item (full update).
    For simplicity, we reuse ItemCreate schema:
    - name, sku, qty, price will be replaced.
    """
    db_item = db.query(models.Item).filter(models.Item.id == item_id).first()
    if not db_item:
        raise HTTPException(status_code=404, detail="Item not found")

    db_item.name = item.name
    db_item.sku = item.sku
    db_item.qty = item.qty
    db_item.price = item.price

    db.commit()
    db.refresh(db_item)
    return db_item


@app.delete("/items/{item_id}")
def delete_one_item(item_id: int, db: Session = Depends(get_db)):
    """
    Deletes an item.
    """
    db_item = db.query(models.Item).filter(models.Item.id == item_id).first()
    if not db_item:
        raise HTTPException(status_code=404, detail="Item not found")

    db.delete(db_item)
    db.commit()
    return {"deleted": True, "id": item_id}


# -----------------------------
# Export Helpers
# -----------------------------
def _export_filename(base: str, ext: str) -> str:
    """
    Creates a nice filename like:
    inventory_2026-01-29_18-32-10.xlsx
    """
    ts = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    return f"{base}_{ts}.{ext}"


# -----------------------------
# Export: Excel (XLSX)
# -----------------------------
@app.get("/export/items.xlsx")
def export_items_xlsx(db: Session = Depends(get_db)):
    """
    Exports all items to an Excel file (modern .xlsx).
    Requires: openpyxl
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    items = db.query(models.Item).order_by(models.Item.id.asc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"

    headers = ["ID", "Name", "SKU", "QTY", "Price", "Value"]
    ws.append(headers)

    # Style header row
    header_font = Font(bold=True)
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for it in items:
        price = it.price if it.price is not None else None
        value = (it.qty * it.price) if (it.price is not None) else None
        ws.append([it.id, it.name, it.sku, it.qty, price, value])

    # Set column widths (simple auto-ish sizing)
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    # Save to bytes
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = _export_filename("inventory", "xlsx")
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# -----------------------------
# Export: PDF (A4)
# -----------------------------
@app.get("/export/items.pdf")
def export_items_pdf(db: Session = Depends(get_db)):
    """
    Exports all items to an A4 PDF.
    Requires: reportlab
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet

    items = db.query(models.Item).order_by(models.Item.id.asc()).all()

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=24, rightMargin=24, topMargin=24, bottomMargin=24)

    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph("Inventory Report (A4)", styles["Title"]))
    story.append(Paragraph(datetime.now().strftime("%Y-%m-%d %H:%M:%S"), styles["Normal"]))
    story.append(Spacer(1, 12))

    data = [["ID", "Name", "SKU", "QTY", "Price", "Value"]]
    for it in items:
        price = "" if it.price is None else f"{it.price:.2f}"
        value = "" if it.price is None else f"{(it.qty * it.price):.2f}"
        data.append([str(it.id), it.name, it.sku, str(it.qty), price, value])

    table = Table(data, colWidths=[40, 170, 120, 60, 60, 70])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F172A")),  # dark header
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )

    story.append(table)
    doc.build(story)

    buffer.seek(0)
    filename = _export_filename("inventory", "pdf")
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
